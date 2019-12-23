import openpyxl
import requests
import jsonpath
import json



def test_Add_():
    global id
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    f = open('C:\\Users\\umair.anis\\API_Automation\\AddStudent.json')
    wk = openpyxl.load_workbook('C:\\Users\\umair.anis\\API_Automation\\multiple_students.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row
    json_request = json.loads(f.read())

    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request['first_name']=cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value
        response = requests.post(API_URL,json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
